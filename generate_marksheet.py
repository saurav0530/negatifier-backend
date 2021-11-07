import os, sys, csv, shutil, PIL
from openpyxl import load_workbook, Workbook, workbook

from openpyxl.styles import alignment
from openpyxl.styles import Font, DEFAULT_FONT
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border,Side
from openpyxl.drawing.image import Image

Img = PIL.Image.open('./logo.png')
Img = Img.resize((898,98))
Img.save('./logo.png')

positive = float(sys.argv[1])
negative = float(sys.argv[2])
side = side = Side(style='thin', color='000000')
thin_border = Border(left=side,right=side,bottom=side,top=side)


DEFAULT_FONT.name = "Century"
DEFAULT_FONT.size = 12
# Checking/Creating output directory
try:
    os.mkdir("./output")
except FileExistsError:
    shutil.rmtree("./output")
    os.mkdir("./output")
    pass

response = list()

answers = list()
with open('./input/responses.csv', 'r') as csvfile:
    reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
    for data in reader:
        if(data[6]=='ANSWER'):
            answers=data
            break

def create_marksheet_for_each_student():
    with open('./input/master_roll.csv', 'r') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
        for data in reader:
            if(data[0]=='roll' or data[0]=='answer'):
                continue

            wb = Workbook()
            ws = wb['Sheet']
            ws.title = 'quiz'

            ws.merge_cells('A5:E5')
            ws.merge_cells('A1:E4')
            ws['A5'] = 'Mark Sheet'
            ws["A5"].alignment = Alignment(horizontal="center", vertical='center')
            ws['A5'].font = Font(name = "Century", size=18,underline='singleAccounting',bold=True,color='000000')
            ws['A5'].border = thin_border
            ws['A1'].border = thin_border

            img = Image('./logo.png')
            img.anchor = 'A1'
            ws.add_image(img)
            
            for i in range(1,41):
                ws.row_dimensions[i].height = (77.25)/4
            ws.row_dimensions[5].height = 40
            
            ws['A6'] = 'Name: '
            ws['D6'] = 'Exam: '
            ws['A7'] = 'Roll Number: '
            ws['B6'] = data[1]
            ws['E6'] = 'Quiz'
            ws['B7'] = data[0]
            ws['B9'] = 'Right'
            ws['C9'] = 'Wrong'
            ws['D9'] = 'Not Attempt'
            ws['E9'] = 'Max'
            ws['A10'] = 'No.'
            ws['A11'] = 'Marking'
            ws['A12'] = 'Total'
            
            cells = ['B9','C9','D9','E9','A10','A11','A12','A9']
            for a in cells:
                ws[a].font = Font(name = "Century",bold=True,color='000000')
                ws[a].alignment = Alignment(horizontal="center", vertical='center')
                ws[a].border = thin_border
            for col in 'ABCDE':
                ws.column_dimensions[col].width = 20

            ws['A15'].alignment = Alignment(horizontal="center", vertical='center')
            ws['B15'].alignment = Alignment(horizontal="center", vertical='center')

            ws['A15'] = 'Student Ans'
            ws['B15'] = 'Correct Ans'
            
            for i in range(7,min(32,len(answers))):
                ws['B'+str(i+9)] = answers[i]
                
                ws['B'+str(i+9)].font = Font(name = "Century",color='0000FF')
                ws['A'+str(i+9)].border = thin_border
                ws['B'+str(i+9)].border = thin_border
                ws['A'+str(i+9)].alignment = Alignment(horizontal="center", vertical='center')
                ws['B'+str(i+9)].alignment = Alignment(horizontal="center", vertical='center')

            ws['D15'] = 'Student Ans'
            ws['E15'] = 'Correct Ans'
            for i in range(32,len(answers)):
                ws['E'+str(i-16)] = answers[i]
                
                ws['E'+str(i-16)].font = Font(name = "Century",color='0000FF')
                ws['D'+str(i-16)].border = thin_border
                ws['E'+str(i-16)].border = thin_border
                ws['D'+str(i-16)].alignment = Alignment(horizontal="center", vertical='center')
                ws['E'+str(i-16)].alignment = Alignment(horizontal="center", vertical='center')

            ws['A15'].border = thin_border
            ws['B15'].border = thin_border
            ws['A15'].font = Font(name = "Century",bold=True,color='000000')
            ws['B15'].font = Font(name = "Century",bold=True,color='000000')

            ws['D15'].alignment = Alignment(horizontal="center", vertical='center')
            ws['E15'].alignment = Alignment(horizontal="center", vertical='center')
            ws['D15'].border = thin_border
            ws['E15'].border = thin_border
            ws['D15'].font = Font(name = "Century",bold=True,color='000000')
            ws['E15'].font = Font(name = "Century",bold=True,color='000000')

            ws['B6'].font = Font(name = "Century",bold=True,color='000000')
            ws['B7'].font = Font(name = "Century",bold=True,color='000000')
            ws['E6'].font = Font(name = "Century",bold=True,color='000000')
            ws['A6'].alignment = Alignment(horizontal='right')
            ws['A7'].alignment = Alignment(horizontal='right')
            ws['D6'].alignment = Alignment(horizontal='right')
            ws['E6'].alignment = Alignment(horizontal='left')
            ws['B6'].alignment = Alignment(horizontal='left')
            ws['B7'].alignment = Alignment(horizontal='left')

            
            ws["A1"].alignment = Alignment(horizontal="center",vertical="center")

            green = ['B10','B11','B12']
            red   = ['C10','C11','C12']
            cells = ['B10','B11','B12','C10','C11','C12','D10','D11','D12','E10','E11','E12']

            for i in green:
                ws[i].font = Font(name = "Century",color='00FF00')
            for i in red:
                ws[i].font = Font(name = "Century",color='FF0000')
            for cell in cells:
                ws[cell].border = thin_border
                ws[cell].alignment = Alignment(horizontal="center", vertical='center')

            ws['E12'].font = Font(name = "Century",color='0000FF')

            wb.save('./output/'+data[0]+'.xlsx')

def add_marks_to_marksheet(data):
    wb = load_workbook('./output/'+data[6]+'.xlsx')
    ws = wb['quiz']
    
    correct=0
    incorrect=0
    left=0
    
    for i in range(7,min(32,len(answers))):
        ws['A'+str(i+9)] = data[i]
        if(data[i] and data[i]==answers[i]):
            ws['A'+str(i+9)].font = Font(name = "Century",color='00FF00')
            correct+=1
        elif(data[i]):
            ws['A'+str(i+9)].font = Font(name = "Century",color='FF0000')
            incorrect+=1
        else:
            left+=1

    for i in range(32,len(answers)):
        ws['D'+str(i-16)] = data[i]
        if(data[i] and data[i]==answers[i]):
            ws['D'+str(i-16)].font = Font(name = "Century",color='00FF00')
            correct+=1
        elif(data[i]):
            ws['D'+str(i-16)].font = Font(name = "Century",color='FF0000')
            incorrect+=1
        else:
            left+=1

    ws['B10']=correct
    ws['C10']=incorrect
    ws['D10']=left
    ws['B11']=sys.argv[1]
    ws['C11']=sys.argv[2]
    ws['D11']=0
    ws['B12']=ws['B10'].value * float(sys.argv[1])
    ws['C12']=ws['C10'].value * float(sys.argv[2])
    ws['E10']=correct+incorrect+left
    numr = ws['B12'].value + ws['C12'].value
    numr = (int(numr*100))/100
    denomr = ws['E10'].value * float(sys.argv[1])
    denomr = (int(denomr*100))/100
    ws['E12']=str(numr)+'/'+str(denomr)
    

    wb.save('./output/'+data[6]+'.xlsx')

def exec_responses():
    with open('./input/responses.csv','r') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
        for data in reader:
            if(data[6]=='Roll Number'):
                continue
            add_marks_to_marksheet(data)
            response.append(data[6])


if(len(answers)==0):
    print('00')
else:
    create_marksheet_for_each_student()
    exec_responses()
    print(len(response))