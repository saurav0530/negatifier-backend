import csv, shutil
from openpyxl import load_workbook, Workbook, workbook

try:
    with open('./output/concise_markheet.csv', 'a', newline='') as file:
        writer_obj = csv.writer(file)
        with open('./input/responses.csv','r') as csvfile:
            reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
            for data in reader:
                roll = data[6]
                if(roll=='Roll Number'):
                    data.insert(6,'Score_After_Negative')
                    data.insert(len(data),'statusAns')
                    writer_obj.writerow(data)
                    continue
                wb = load_workbook('./output/'+roll+'.xlsx')
                ws = wb['quiz']
                data.insert(6,ws['E12'].value)
                data.insert(len(data),'['+str(ws['B10'].value)+','+str(ws['C10'].value)+','+str(ws['D10'].value)+']')
                wb.close()
                writer_obj.writerow(data)
        file.close()
    shutil.make_archive('./marksheet','zip','./output')
except Exception as e:
    print(e)